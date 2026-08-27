#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
F.R.I.D.A.Y. — File Retrieval, Indexing, Directory & Archiving Y-system
Núcleo de organização inteligente de arquivos e projetos.

Filosofia: Understand first. Organize second. Delete never.

by Axolotl BR
"""

import csv
import hashlib
import json
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

VERSAO = "0.2"
NOME_PASTA_DADOS = ".friday"
ARQUIVO_REGRAS = "friday_regras.json"
ARQUIVO_CATALOGO = "catalogo.json"
PASTA_LIXEIRA = "_Lixeira"

# --------------------------------------------------------------------------- #
# Categorias por extensão
# --------------------------------------------------------------------------- #
CATEGORIAS_EXT = {
    "Documentos":    [".doc", ".docx", ".odt", ".txt", ".rtf", ".md", ".pages"],
    "Planilhas":     [".xls", ".xlsx", ".ods", ".csv"],
    "Apresentacoes": [".ppt", ".pptx", ".odp", ".key"],
    "PDFs":          [".pdf"],
    "Imagens":       [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".heic",
                       ".svg", ".tiff", ".psd", ".ai", ".xd", ".fig"],
    "Videos":        [".mp4", ".mov", ".avi", ".mkv", ".wmv", ".flv", ".webm"],
    "Musica":        [".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a"],
    "Compactados":   [".zip", ".rar", ".7z", ".tar", ".gz", ".bz2"],
    "Programas":     [".exe", ".msi", ".dmg", ".pkg", ".deb", ".appimage"],
    "Codigo":        [".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css", ".java",
                       ".c", ".cpp", ".json", ".xml", ".sh", ".rb", ".go", ".rs"],
}

EXT_PROJETO_EDICAO = {".prproj", ".aep", ".veg", ".drp", ".fcpxml"}
MARCADORES_PROJETO_CODIGO = {".git", "package.json", "requirements.txt", "Cargo.toml",
                              "CMakeLists.txt", "pyproject.toml", "go.mod"}

PALAVRAS_TIPO_VIDEO = {
    "Clips":      ["clip", "clipe"],
    "Highlights": ["highlight", "melhores momentos", "best of", "best"],
    "Gameplay":   ["gameplay", "jogando", "play"],
    "Recordings": ["gravacao", "gravação", "recording", "record"],
}

# Arquivos e extensões considerados "lixo" — nunca essenciais, seguros de mover
NOMES_LIXO = {"thumbs.db", "desktop.ini", ".ds_store", "ehthumbs.db", ".spotlight-v100",
              ".trashes"}
EXT_LIXO = {".tmp", ".temp", ".bak", ".crdownload", ".part", ".download", ".log"}


def regras_padrao():
    return {
        "modo": "SMART",  # SAFE | SMART | AUTO
        "pastas_protegidas": ["Downloads", "node_modules", ".git", "AppData", "System32"],
        "jogos": ["CS2", "Counter-Strike", "Deadlock", "Minecraft", "Valorant",
                  "League of Legends", "LOL", "Apex Legends", "Fortnite", "GTA V", "GTA",
                  "Rocket League", "Among Us", "Overwatch"],
        # regra manual: {"palavra_chave": "deadlock", "destino": "VIDEOS/GAMES/Deadlock/Clips"}
        "regras_personalizadas": [],
        # aprendidas a partir de correções do usuário: mesmo formato
        "regras_aprendidas": [],
    }


class GerenciadorRegras:
    """Carrega/salva as regras do usuário (friday_regras.json) na pasta raiz organizada."""

    def __init__(self, pasta_raiz: Path):
        self.pasta_raiz = Path(pasta_raiz)
        self.pasta_dados = self.pasta_raiz / NOME_PASTA_DADOS
        self.caminho = self.pasta_dados / ARQUIVO_REGRAS
        self.regras = self._carregar()

    def _carregar(self):
        if self.caminho.exists():
            try:
                with open(self.caminho, encoding="utf-8") as f:
                    dados = json.load(f)
                base = regras_padrao()
                base.update(dados)
                return base
            except (json.JSONDecodeError, OSError):
                pass
        return regras_padrao()

    def salvar(self):
        self.pasta_dados.mkdir(exist_ok=True)
        with open(self.caminho, "w", encoding="utf-8") as f:
            json.dump(self.regras, f, ensure_ascii=False, indent=2)

    def adicionar_regra_personalizada(self, palavra_chave, destino):
        self.regras["regras_personalizadas"].append(
            {"palavra_chave": palavra_chave.lower(), "destino": destino}
        )
        self.salvar()

    def aprender(self, palavra_chave, destino):
        self.regras["regras_aprendidas"].append(
            {"palavra_chave": palavra_chave.lower(), "destino": destino}
        )
        self.salvar()

    def esta_protegida(self, caminho: Path) -> bool:
        partes = {p.lower() for p in caminho.parts}
        for protegida in self.regras["pastas_protegidas"]:
            if protegida.lower() in partes:
                return True
        return False


class Catalogo:
    """Índice persistente de tudo que o F.R.I.D.A.Y. já organizou, com ID e tags
    próprias — a base para a planilha exportável e para buscas futuras."""

    def __init__(self, pasta_raiz: Path):
        self.pasta_raiz = Path(pasta_raiz)
        self.pasta_dados = self.pasta_raiz / NOME_PASTA_DADOS
        self.caminho = self.pasta_dados / ARQUIVO_CATALOGO
        self._dados = self._carregar()

    def _carregar(self):
        if self.caminho.exists():
            try:
                with open(self.caminho, encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                pass
        return {"proximo_id": 1, "itens": {}}

    def salvar(self):
        self.pasta_dados.mkdir(exist_ok=True)
        with open(self.caminho, "w", encoding="utf-8") as f:
            json.dump(self._dados, f, ensure_ascii=False, indent=2)

    def novo_id(self) -> str:
        n = self._dados["proximo_id"]
        self._dados["proximo_id"] = n + 1
        return f"FRD-{n:05d}"

    def adicionar(self, nome, caminho, categoria, projeto, tags, tamanho, hash_):
        id_ = self.novo_id()
        self._dados["itens"][id_] = {
            "id": id_,
            "nome": nome,
            "caminho": str(caminho),
            "categoria": categoria,
            "projeto": projeto or "",
            "tags": tags,
            "tamanho": tamanho,
            "hash": hash_ or "",
            "adicionado_em": datetime.now().isoformat(timespec="seconds"),
        }
        self.salvar()
        return id_

    def remover(self, id_):
        self._dados["itens"].pop(id_, None)
        self.salvar()

    def atualizar_caminho(self, id_, novo_caminho):
        if id_ in self._dados["itens"]:
            self._dados["itens"][id_]["caminho"] = str(novo_caminho)
            self.salvar()

    def listar(self):
        return list(self._dados["itens"].values())


# --------------------------------------------------------------------------- #
# Utilidades
# --------------------------------------------------------------------------- #
def tamanho_legivel(n_bytes):
    n = float(n_bytes)
    for unidade in ["B", "KB", "MB", "GB"]:
        if n < 1024:
            return f"{n:.0f} {unidade}" if unidade == "B" else f"{n:.1f} {unidade}"
        n /= 1024
    return f"{n:.1f} TB"


def hash_arquivo(caminho: Path, bloco=65536, limite_bytes=200 * 1024 * 1024):
    try:
        if caminho.stat().st_size > limite_bytes:
            return None
        h = hashlib.md5()
        with open(caminho, "rb") as f:
            while chunk := f.read(bloco):
                h.update(chunk)
        return h.hexdigest()
    except (OSError, PermissionError):
        return None


def caminho_sem_colisao(destino: Path) -> Path:
    if not destino.exists():
        return destino
    base, ext = destino.stem, destino.suffix
    i = 1
    while True:
        candidato = destino.with_name(f"{base} ({i}){ext}")
        if not candidato.exists():
            return candidato
        i += 1


# --------------------------------------------------------------------------- #
# Estruturas
# --------------------------------------------------------------------------- #
@dataclass
class ItemPlanejado:
    origem: Path
    destino: Path
    categoria: str
    motivo: str
    confianca: float          # 0.0 - 1.0
    tamanho: int
    e_pasta: bool = False
    duplicado: bool = False
    projeto: Optional[str] = None
    tags: list = field(default_factory=list)


@dataclass
class ItemLimpeza:
    caminho: Path
    tipo: str          # "pasta_vazia" | "arquivo_lixo"
    motivo: str
    tamanho: int = 0


# --------------------------------------------------------------------------- #
# Classificação
# --------------------------------------------------------------------------- #
def categoria_por_extensao(ext: str) -> str:
    ext = ext.lower()
    for nome, extensoes in CATEGORIAS_EXT.items():
        if ext in extensoes:
            return nome
    return "Outros"


def detectar_jogo(caminho: Path, jogos: list) -> Optional[str]:
    alvo = f"{caminho.parent.name} {caminho.stem}".lower()
    for jogo in jogos:
        if jogo.lower() in alvo:
            return jogo
    return None


def detectar_tipo_video(caminho: Path) -> str:
    alvo = f"{caminho.parent.name} {caminho.stem}".lower()
    for tipo, palavras in PALAVRAS_TIPO_VIDEO.items():
        if any(p in alvo for p in palavras):
            return tipo
    return "Recordings"


def pasta_e_projeto_codigo(pasta: Path) -> bool:
    try:
        nomes = {p.name for p in pasta.iterdir()}
    except (OSError, PermissionError):
        return False
    return bool(nomes & MARCADORES_PROJETO_CODIGO)


def pasta_e_projeto_edicao(pasta: Path) -> bool:
    try:
        for p in pasta.iterdir():
            if p.suffix.lower() in EXT_PROJETO_EDICAO:
                return True
    except (OSError, PermissionError):
        return False
    return False


def gerar_tags(item: ItemPlanejado, ext: str = "") -> list:
    tags = {item.categoria.lower()}
    if item.projeto:
        tags.add(item.projeto.lower())
    if ext:
        tags.add(ext.lstrip(".").lower())
    if item.duplicado:
        tags.add("duplicado")
    try:
        ano = datetime.fromtimestamp(item.origem.stat().st_mtime).strftime("%Y")
        tags.add(ano)
    except OSError:
        pass
    return sorted(t for t in tags if t)


def classificar_regra_manual(caminho: Path, regras: dict) -> Optional[ItemPlanejado]:
    """Regras do usuário (personalizadas + aprendidas) têm prioridade máxima."""
    alvo = f"{caminho.parent.name} {caminho.name}".lower()
    todas = regras["regras_personalizadas"] + regras["regras_aprendidas"]
    for regra in todas:
        if regra["palavra_chave"] in alvo:
            return ItemPlanejado(
                origem=caminho,
                destino=Path(regra["destino"]) / caminho.name,
                categoria="Regra do usuário",
                motivo=f"corresponde à regra \"{regra['palavra_chave']}\"",
                confianca=1.0,
                tamanho=caminho.stat().st_size if caminho.is_file() else 0,
                projeto=regra["palavra_chave"],
            )
    return None


def classificar_arquivo(caminho: Path, regras: dict) -> ItemPlanejado:
    tamanho = caminho.stat().st_size

    manual = classificar_regra_manual(caminho, regras)
    if manual:
        manual.tags = gerar_tags(manual, caminho.suffix)
        return manual

    ext = caminho.suffix.lower()

    if ext in CATEGORIAS_EXT["Videos"]:
        jogo = detectar_jogo(caminho, regras["jogos"])
        if jogo:
            tipo = detectar_tipo_video(caminho)
            item = ItemPlanejado(
                origem=caminho,
                destino=Path("VIDEOS") / "GAMES" / jogo / tipo / caminho.name,
                categoria="Vídeo de jogo",
                motivo=f"jogo \"{jogo}\" identificado pelo nome/pasta",
                confianca=0.85,
                tamanho=tamanho,
                projeto=jogo,
            )
            item.tags = gerar_tags(item, ext) + [tipo.lower()]
            return item
        item = ItemPlanejado(
            origem=caminho,
            destino=Path("UNSORTED") / caminho.name,
            categoria="Vídeo (não identificado)",
            motivo="não foi possível identificar o jogo com confiança",
            confianca=0.3,
            tamanho=tamanho,
        )
        item.tags = gerar_tags(item, ext)
        return item

    categoria = categoria_por_extensao(ext)
    item = ItemPlanejado(
        origem=caminho,
        destino=Path(categoria) / caminho.name,
        categoria=categoria,
        motivo=f"extensão {ext or '(sem extensão)'}",
        confianca=0.8 if categoria != "Outros" else 0.4,
        tamanho=tamanho,
    )
    item.tags = gerar_tags(item, ext)
    return item


def classificar_pasta(pasta: Path) -> Optional[ItemPlanejado]:
    """Detecta se uma subpasta é um projeto (código ou edição) e deve ser
    movida inteira, sem mexer no conteúdo interno."""
    tamanho = sum(f.stat().st_size for f in pasta.rglob("*") if f.is_file())

    if pasta_e_projeto_codigo(pasta):
        item = ItemPlanejado(
            origem=pasta, destino=Path("PROJECTS") / "CODE" / pasta.name,
            categoria="Projeto de código", motivo="contém marcadores de projeto (.git, package.json, etc.)",
            confianca=0.9, tamanho=tamanho, e_pasta=True, projeto=pasta.name,
        )
        item.tags = gerar_tags(item)
        return item

    if pasta_e_projeto_edicao(pasta):
        item = ItemPlanejado(
            origem=pasta, destino=Path("PROJECTS") / "EDITING" / pasta.name,
            categoria="Projeto de edição", motivo="contém arquivos de projeto de edição (.prproj, .aep, etc.)",
            confianca=0.9, tamanho=tamanho, e_pasta=True, projeto=pasta.name,
        )
        item.tags = gerar_tags(item)
        return item

    return None


# --------------------------------------------------------------------------- #
# Pastas vazias e arquivos "lixo" (limpeza)
# --------------------------------------------------------------------------- #
def e_arquivo_lixo(caminho: Path) -> bool:
    if caminho.name.lower() in NOMES_LIXO:
        return True
    return caminho.suffix.lower() in EXT_LIXO


def _pasta_esta_vazia(pasta: Path) -> bool:
    """Vazia de verdade: sem nenhum arquivo em nenhum nível abaixo dela."""
    try:
        for item in pasta.rglob("*"):
            if item.is_file():
                return False
        return True
    except (OSError, PermissionError):
        return False


def encontrar_pastas_vazias(raiz: Path, protegidas_check) -> list:
    vazias = []
    for pasta in sorted(raiz.rglob("*"), reverse=True):  # de dentro para fora
        if not pasta.is_dir():
            continue
        if pasta.name.startswith(".") or pasta.name == NOME_PASTA_DADOS:
            continue
        if protegidas_check(pasta):
            continue
        if _pasta_esta_vazia(pasta):
            vazias.append(pasta)
    return vazias


def encontrar_arquivos_lixo(raiz: Path, protegidas_check) -> list:
    achados = []
    for caminho in raiz.rglob("*"):
        if not caminho.is_file():
            continue
        if protegidas_check(caminho):
            continue
        if e_arquivo_lixo(caminho):
            achados.append(caminho)
    return achados


# --------------------------------------------------------------------------- #
# Motor principal
# --------------------------------------------------------------------------- #
PASTAS_RESERVADAS = set(CATEGORIAS_EXT.keys()) | {
    "VIDEOS", "PROJECTS", "UNSORTED", "_Duplicados", PASTA_LIXEIRA, NOME_PASTA_DADOS,
}


class Friday:
    def __init__(self, pasta_raiz):
        self.pasta_raiz = Path(pasta_raiz).expanduser().resolve()
        self.regras_mgr = GerenciadorRegras(self.pasta_raiz)
        self.catalogo = Catalogo(self.pasta_raiz)

    # ---- leitura ---- #
    def _itens_nivel_1(self):
        for entrada in self.pasta_raiz.iterdir():
            if entrada.name.startswith("."):
                continue
            if entrada.is_dir() and entrada.name in PASTAS_RESERVADAS:
                continue
            yield entrada

    def _encontrar_duplicados(self, arquivos):
        hashes = {}
        for caminho in arquivos:
            h = hash_arquivo(caminho)
            if h:
                hashes.setdefault(h, []).append(caminho)
        return {h: v for h, v in hashes.items() if len(v) > 1}

    # ---- planejamento: organização ---- #
    def montar_plano(self):
        regras = self.regras_mgr.regras
        itens = list(self._itens_nivel_1())
        arquivos = [i for i in itens if i.is_file() and not e_arquivo_lixo(i)]
        pastas = [i for i in itens if i.is_dir()]

        duplicados = self._encontrar_duplicados(arquivos)
        ids_duplicados_secundarios = set()
        for grupo in duplicados.values():
            ordenado = sorted(grupo, key=lambda c: c.stat().st_mtime)
            for c in ordenado[1:]:
                ids_duplicados_secundarios.add(str(c))

        plano = []

        for caminho in arquivos:
            if self.regras_mgr.esta_protegida(caminho):
                continue
            if str(caminho) in ids_duplicados_secundarios:
                item = ItemPlanejado(
                    origem=caminho, destino=Path("_Duplicados") / caminho.name,
                    categoria="Duplicado", motivo="cópia idêntica de outro arquivo já existente",
                    confianca=0.95, tamanho=caminho.stat().st_size, duplicado=True,
                )
                item.tags = gerar_tags(item, caminho.suffix)
                plano.append(item)
                continue
            plano.append(classificar_arquivo(caminho, regras))

        for pasta in pastas:
            if self.regras_mgr.esta_protegida(pasta):
                continue
            item = classificar_pasta(pasta)
            if item:
                plano.append(item)
            # pastas que não são projetos reconhecidos ficam como estão
            # (F.R.I.D.A.Y. nunca reorganiza o conteúdo interno sem confiança)

        return plano, duplicados

    # ---- planejamento: limpeza ---- #
    def montar_plano_limpeza(self):
        protegidas_check = self.regras_mgr.esta_protegida
        itens = []
        for caminho in encontrar_arquivos_lixo(self.pasta_raiz, protegidas_check):
            if caminho.parent.name == PASTA_LIXEIRA:
                continue
            itens.append(ItemLimpeza(caminho=caminho, tipo="arquivo_lixo",
                                      motivo="arquivo temporário/lixo do sistema",
                                      tamanho=caminho.stat().st_size))
        for pasta in encontrar_pastas_vazias(self.pasta_raiz, protegidas_check):
            itens.append(ItemLimpeza(caminho=pasta, tipo="pasta_vazia",
                                      motivo="pasta sem nenhum arquivo dentro"))
        return itens

    def aplicar_limpeza(self, itens):
        registro = []
        erros = []
        for item in itens:
            if item.tipo == "arquivo_lixo":
                destino = self.pasta_raiz / PASTA_LIXEIRA / item.caminho.relative_to(self.pasta_raiz).name
                destino.parent.mkdir(parents=True, exist_ok=True)
                destino = caminho_sem_colisao(destino)
                try:
                    shutil.move(str(item.caminho), str(destino))
                    registro.append({"tipo": "movido", "origem": str(item.caminho),
                                      "destino": str(destino), "categoria": "Limpeza",
                                      "e_pasta": False})
                except (OSError, PermissionError) as e:
                    erros.append((item.caminho.name, str(e)))
            elif item.tipo == "pasta_vazia":
                try:
                    if _pasta_esta_vazia(item.caminho) and item.caminho.exists():
                        item.caminho.rmdir()
                        registro.append({"tipo": "pasta_removida", "origem": str(item.caminho),
                                          "destino": None, "categoria": "Limpeza", "e_pasta": True})
                except (OSError, PermissionError) as e:
                    erros.append((item.caminho.name, str(e)))

        if registro:
            self._salvar_operacao(registro)
        return registro, erros

    # ---- aplicação: organização ---- #
    def aplicar(self, plano, aprovados=None):
        """aprovados: conjunto opcional de índices do plano a aplicar (modo preview).
        Se None, aplica tudo."""
        if aprovados is not None:
            plano = [item for i, item in enumerate(plano) if i in aprovados]

        registro = []
        erros = []
        for item in plano:
            destino_final = self.pasta_raiz / item.destino
            destino_final.parent.mkdir(parents=True, exist_ok=True)
            destino_final = caminho_sem_colisao(destino_final)
            try:
                shutil.move(str(item.origem), str(destino_final))
            except (OSError, PermissionError) as e:
                erros.append((item.origem.name, str(e)))
                continue

            h = hash_arquivo(destino_final) if not item.e_pasta else None
            catalogo_id = self.catalogo.adicionar(
                nome=destino_final.name, caminho=destino_final, categoria=item.categoria,
                projeto=item.projeto, tags=item.tags, tamanho=item.tamanho, hash_=h,
            )
            registro.append({
                "tipo": "movido",
                "origem": str(item.origem),
                "destino": str(destino_final),
                "categoria": item.categoria,
                "e_pasta": item.e_pasta,
                "catalogo_id": catalogo_id,
            })

        if registro:
            self._salvar_operacao(registro)
        return registro, erros

    def _salvar_operacao(self, registro):
        pasta_dados = self.pasta_raiz / NOME_PASTA_DADOS
        pasta_dados.mkdir(exist_ok=True)
        op_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        log_path = pasta_dados / f"operacao_{op_id}.json"
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump({
                "id": op_id,
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "itens": registro,
            }, f, ensure_ascii=False, indent=2)

    # ---- histórico / undo ---- #
    def historico(self):
        pasta_dados = self.pasta_raiz / NOME_PASTA_DADOS
        if not pasta_dados.exists():
            return []
        operacoes = []
        for arq in sorted(pasta_dados.glob("operacao_*.json"), reverse=True):
            try:
                with open(arq, encoding="utf-8") as f:
                    operacoes.append(json.load(f))
            except (json.JSONDecodeError, OSError):
                continue
        return operacoes

    def desfazer(self, operacao_id):
        pasta_dados = self.pasta_raiz / NOME_PASTA_DADOS
        log_path = pasta_dados / f"operacao_{operacao_id}.json"
        if not log_path.exists():
            return 0, ["operação não encontrada"]

        with open(log_path, encoding="utf-8") as f:
            dados = json.load(f)

        restaurados = 0
        erros = []
        for mov in reversed(dados["itens"]):
            tipo = mov.get("tipo", "movido")
            if tipo == "pasta_removida":
                try:
                    Path(mov["origem"]).mkdir(parents=True, exist_ok=True)
                    restaurados += 1
                except OSError as e:
                    erros.append((Path(mov["origem"]).name, str(e)))
                continue

            origem, destino = Path(mov["origem"]), Path(mov["destino"])
            if destino.exists():
                origem.parent.mkdir(parents=True, exist_ok=True)
                volta = caminho_sem_colisao(origem)
                try:
                    shutil.move(str(destino), str(volta))
                    restaurados += 1
                    if mov.get("catalogo_id"):
                        self.catalogo.remover(mov["catalogo_id"])
                except (OSError, PermissionError) as e:
                    erros.append((destino.name, str(e)))
            else:
                erros.append((destino.name, "arquivo não encontrado no destino"))

        log_path.unlink()
        return restaurados, erros

    # ---- planilha ---- #
    def gerar_planilha(self, caminho_saida: Optional[Path] = None):
        """Gera uma planilha (xlsx se a biblioteca openpyxl estiver disponível,
        senão csv) com o catálogo completo: ID, nome, categoria, projeto, tags,
        tamanho, caminho e data. Retorna (caminho_gerado, formato)."""
        itens = self.catalogo.listar()
        colunas = ["id", "nome", "categoria", "projeto", "tags", "tamanho_legivel",
                   "tamanho_bytes", "caminho", "hash", "adicionado_em"]

        linhas = []
        for it in sorted(itens, key=lambda x: x["id"]):
            linhas.append([
                it["id"], it["nome"], it["categoria"], it.get("projeto", ""),
                ", ".join(it.get("tags", [])), tamanho_legivel(it["tamanho"]),
                it["tamanho"], it["caminho"], it.get("hash", ""), it["adicionado_em"],
            ])

        if caminho_saida is None:
            pasta_dados = self.pasta_raiz / NOME_PASTA_DADOS
            pasta_dados.mkdir(exist_ok=True)
            caminho_saida = pasta_dados / "friday_catalogo"

        caminho_saida = Path(caminho_saida)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Catálogo F.R.I.D.A.Y."
            cabecalhos = ["ID", "Nome", "Categoria", "Projeto/Jogo", "Tags",
                          "Tamanho", "Bytes", "Caminho", "Hash", "Adicionado em"]
            ws.append(cabecalhos)
            for cel in ws[1]:
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = PatternFill("solid", fgColor="E06C82")
            for linha in linhas:
                ws.append(linha)
            larguras = [10, 30, 18, 18, 24, 12, 10, 50, 20, 20]
            for i, largura in enumerate(larguras, start=1):
                ws.column_dimensions[get_column_letter(i)].width = largura
            ws.freeze_panes = "A2"
            if linhas:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{len(linhas) + 1}"

            destino = caminho_saida.with_suffix(".xlsx")
            wb.save(destino)
            return destino, "xlsx"
        except ImportError:
            destino = caminho_saida.with_suffix(".csv")
            with open(destino, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Nome", "Categoria", "Projeto/Jogo", "Tags",
                                  "Tamanho", "Bytes", "Caminho", "Hash", "Adicionado em"])
                writer.writerows(linhas)
            return destino, "csv"
